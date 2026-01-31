# part2_xlsx.py
# 파이썬으로 엑셀 조작하기
# 엑셀을 조작하려면 openpyxl 라이브러리가 필요하다
# 옵션으로 pandas, matplotlib 등을 활용하면 데이터 분석과 시각화도 가능하다.
# pip install openpyxl pandas matplotlib
# 위 명령어를 터미널에 입력하면 해당 환경에 설치된다.

# 엑셀을 다루기 위한 라이브러리 임포트
import openpyxl as oxl # as로 alias(별칭) 설정
from openpyxl import (
    Workbook,
    load_workbook
)


# 엑셀로 데이터 내보내기 함수
def write_to_xlsx(file_path:str):
    # 엑셀 파일 생성
    # 엑셀은 워크북이라는 단위로 파일이 생성된다.
    # wb = oxl.Workbook()
    wb = Workbook() # import에서 해당 클래스, 변수, 함수 등을 직접 임포트한 경우
    # 파일명이나 라이브러리명을 작성하지 않고 직접 사용할 수 있다.

    # 해당 워크북에서 현재 활성화된 시트를 가져온다.
    sheet = wb.active
    # 해당 시트는 표 형태로 되어 있으며, 열은 대문자 알파벳으로 인덱싱하고
    # 행은 숫자1부터 증가하는 형태로 인덱싱되어 있다.
    sheet["A1"] = "이름"
    sheet["B1"] = "나이"
    for i, e in enumerate(["홍길동", "이순신", "장보고"], 2):
        sheet[f"A{i}"] = e
    
    # .save(파일명)
    # 저장하기
    wb.save(file_path)

# 엑셀에서 데이터 읽어오기
def read_from_xlsx(file_path:str):
    wb = load_workbook(file_path)
    sheet = wb.active
    print(sheet["A1"].value)
    for e in range(ord("A"), ord("B")+1):
        for i in range(1, 5):
            print(sheet[f"{chr(e)}{i}"].value)
    
    print([sheet.cell(x, y).value for x in range(1, 10) for y in range(1, 10)])

if __name__ == "__main__":
    file_path = "people.xlsx"
    write_to_xlsx(file_path)
    read_from_xlsx(file_path)